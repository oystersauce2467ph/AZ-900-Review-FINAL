# -*- coding: utf-8 -*-
"""
Generate one image per multiple-choice question for the 5 Kahoot MC files.

WHY: Kahoot's spreadsheet importer cannot carry images - they must be added in
the editor. This script creates a ready-to-drag image per question (themed to
the question's subject area) plus an Image Guide workbook that maps every
question to its image file AND a Kahoot image-library search keyword.

The images are generated (no third-party photos) so there are no licensing
concerns. Each image reflects the SUBJECT AREA of the question/answer (e.g.,
Storage, Networking, Security) rather than printing the literal answer, so it
won't spoil the answer on screen.

Output:
  kahoot_images/Part1..5/PartX_Qnn.png
  AZ-900_Kahoot_Image_Guide.xlsx
"""
import os, re
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Reuse the exact MC question list from the builder
import importlib.util
spec = importlib.util.spec_from_file_location("bk", os.path.join(os.path.dirname(__file__), "build_kahoot_xlsx.py"))
# build script runs on import (writes xlsx) - acceptable; we only need MC
import sys
sys.argv = ["bk"]
bk = importlib.util.module_from_spec(spec)
spec.loader.exec_module(bk)
MC = bk.MC
TF = bk.TF

W, H = 1280, 720
FONT_PATH = "/usr/share/fonts/google-noto-vf/NotoSans[wght].ttf"

def font(size, weight=400):
    f = ImageFont.truetype(FONT_PATH, size)
    try:
        f.set_variation_by_axes([weight])
    except Exception:
        pass
    return f

# ---- category -> palette ----
CATS = {
    "storage":    ("#0B5394", "#0A3D62", "Storage"),
    "networking": ("#1F6FB2", "#10456E", "Networking"),
    "compute":    ("#107C41", "#0B5630", "Compute"),
    "identity":   ("#6B3FA0", "#46256B", "Identity & Access"),
    "security":   ("#A4262C", "#6E1A1E", "Security"),
    "cost":       ("#B7791F", "#7C5210", "Cost & Pricing"),
    "governance": ("#2B6777", "#194750", "Governance & Mgmt"),
    "cloud":      ("#0078D4", "#004C87", "Cloud Concepts"),
}

def classify(q, ans):
    t = (q + " " + ans).lower()
    def has(*words): return any(w in t for w in words)
    if has("hybrid benefit", "capex", "opex", "pricing", "price", "spot", "reserv",
           "savings", "tco", "budget", "sla", "support plan", "subscription",
           "enterprise agreement", "billing", "cost", "rpo", "rto"):
        return "cost"
    if has("blob", "storage account", "disk", "azure files", "queue", "table",
           "lrs", "zrs", "grs", "gzrs", "ra-grs", "archive", " hot", " cool", " cold",
           "data box", "azcopy", "redundan", "tier", "storage explorer", "file sync"):
        return "storage"
    if has("vnet", "virtual network", "expressroute", "vpn", "dns", "gateway",
           "load balanc", "subnet", "peering", "nsg", "network security", "firewall",
           "application gateway", "alias record", "appliance"):
        return "networking"
    if has("entra", "authenticat", "authoriz", " mfa", "multi-factor", "sso",
           "single sign", "rbac", "identity", "conditional access", "passwordless",
           "role"):
        return "identity"
    if has("defender", "key vault", "ddos", "zero trust", "defense in depth",
           "security", "encryption", "threat", "shared responsib"):
        return "security"
    if has("policy", "initiative", "blueprint", "lock", " tag", "purview",
           "resource manager", "arm template", "advisor", "monitor", "service health",
           "log analytics", "application insights", "governance", "compliance",
           "management group", "resource group", "cloud shell", "powershell", "cli",
           "service trust", "json"):
        return "governance"
    if has("virtual machine", " vm", "scale set", "function", "app service",
           "kubernetes", "aks", "container", "virtual desktop", "iot", "openai",
           "serverless", "compute", "availability set", "update domain", "fault domain"):
        return "compute"
    return "cloud"

def hexc(h):
    h = h.lstrip("#")
    if len(h) == 3:
        h = "".join(ch*2 for ch in h)
    h = h[:6]
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

def rounded(draw, box, r, fill):
    draw.rounded_rectangle(box, radius=r, fill=fill)

# ---- simple shape icons (white) drawn in a box centred at (cx,cy) size s ----
def icon_cloud(d, cx, cy, s, col):
    d.ellipse([cx-s*0.5, cy-s*0.18, cx-s*0.05, cy+s*0.28], fill=col)
    d.ellipse([cx-s*0.22, cy-s*0.38, cx+s*0.28, cy+s*0.18], fill=col)
    d.ellipse([cx+s*0.02, cy-s*0.18, cx+s*0.5, cy+s*0.28], fill=col)
    d.rounded_rectangle([cx-s*0.5, cy+s*0.02, cx+s*0.5, cy+s*0.28], radius=s*0.13, fill=col)

def icon_compute(d, cx, cy, s, col):
    d.rounded_rectangle([cx-s*0.45, cy-s*0.4, cx+s*0.45, cy+s*0.18], radius=s*0.06, fill=col)
    d.rounded_rectangle([cx-s*0.35, cy-s*0.3, cx+s*0.35, cy+s*0.08], radius=s*0.04, fill=hexc("#222"))
    d.rectangle([cx-s*0.12, cy+s*0.18, cx+s*0.12, cy+s*0.34], fill=col)
    d.rounded_rectangle([cx-s*0.3, cy+s*0.34, cx+s*0.3, cy+s*0.44], radius=s*0.05, fill=col)

def icon_storage(d, cx, cy, s, col):
    for i, dy in enumerate((-0.3, -0.02, 0.26)):
        y = cy + s*dy
        d.ellipse([cx-s*0.42, y-s*0.12, cx+s*0.42, y+s*0.12], fill=col)
        d.rectangle([cx-s*0.42, y, cx+s*0.42, y+s*0.18], fill=col)
        d.ellipse([cx-s*0.42, y+s*0.06, cx+s*0.42, y+s*0.3], fill=col)

def icon_network(d, cx, cy, s, col):
    pts = [(cx, cy-s*0.34), (cx-s*0.38, cy+s*0.28), (cx+s*0.38, cy+s*0.28)]
    for p in pts:
        d.line([cx, cy, p[0], p[1]], fill=col, width=int(s*0.06))
    for p in [(cx, cy)] + pts:
        d.ellipse([p[0]-s*0.12, p[1]-s*0.12, p[0]+s*0.12, p[1]+s*0.12], fill=col)

def icon_shield(d, cx, cy, s, col):
    pts = [(cx, cy-s*0.42), (cx+s*0.38, cy-s*0.24), (cx+s*0.38, cy+s*0.1),
           (cx, cy+s*0.44), (cx-s*0.38, cy+s*0.1), (cx-s*0.38, cy-s*0.24)]
    d.polygon(pts, fill=col)
    d.line([cx-s*0.12, cy+s*0.02, cx-s*0.02, cy+s*0.16], fill=hexc("#222"), width=int(s*0.05))
    d.line([cx-s*0.02, cy+s*0.16, cx+s*0.18, cy-s*0.12], fill=hexc("#222"), width=int(s*0.05))

def icon_person(d, cx, cy, s, col):
    d.ellipse([cx-s*0.18, cy-s*0.4, cx+s*0.18, cy-s*0.04], fill=col)
    d.pieslice([cx-s*0.34, cy-s*0.02, cx+s*0.34, cy+s*0.6], 180, 360, fill=col)

def icon_coin(d, cx, cy, s, col):
    d.ellipse([cx-s*0.38, cy-s*0.38, cx+s*0.38, cy+s*0.38], fill=col)
    d.ellipse([cx-s*0.3, cy-s*0.3, cx+s*0.3, cy+s*0.3], outline=hexc("#222"), width=int(s*0.04))
    f = font(int(s*0.55), 700)
    d.text((cx, cy), "$", font=f, fill=hexc("#222"), anchor="mm")

def icon_gear(d, cx, cy, s, col):
    import math
    R, r = s*0.4, s*0.18
    teeth = 8
    pts = []
    for i in range(teeth*2):
        ang = math.pi*i/teeth
        rad = R if i % 2 == 0 else R*0.78
        pts.append((cx+rad*math.cos(ang), cy+rad*math.sin(ang)))
    d.polygon(pts, fill=col)
    d.ellipse([cx-r, cy-r, cx+r, cy+r], fill=hexc("#222"))

ICONS = {
    "storage": icon_storage, "networking": icon_network, "compute": icon_compute,
    "identity": icon_person, "security": icon_shield, "cost": icon_coin,
    "governance": icon_gear, "cloud": icon_cloud,
}

def make_image(path, cat, footer):
    c1, c2, title = CATS[cat]
    img = Image.new("RGB", (W, H), hexc(c2))
    d = ImageDraw.Draw(img)
    # diagonal band
    d.polygon([(0, 0), (W, 0), (W, H*0.55), (0, H*0.78)], fill=hexc(c1))
    # white card
    cardb = [W*0.16, H*0.16, W*0.84, H*0.8]
    # soft shadow
    d.rounded_rectangle([cardb[0]+10, cardb[1]+14, cardb[2]+10, cardb[3]+14], radius=36, fill=hexc("#C9C9C9"))
    rounded(d, cardb, 36, (255, 255, 255))
    # icon circle
    ccx, ccy, cr = W*0.5, H*0.4, 150
    d.ellipse([ccx-cr, ccy-cr, ccx+cr, ccy+cr], fill=hexc(c1))
    ICONS[cat](d, ccx, ccy, cr*1.6, (255, 255, 255))
    # title
    d.text((W*0.5, H*0.64), title, font=font(58, 700), fill=hexc(c2), anchor="mm")
    # footer
    d.text((W*0.5, H*0.73), footer, font=font(30, 500), fill=hexc("#555555"), anchor="mm")
    img.save(path, "PNG")

def keyword(cat, ans):
    a = re.sub(r"\(.*?\)", "", ans).strip()
    # if answer is numeric/boolean/format token, fall back to subject area
    if re.fullmatch(r"[0-9.,]+|yes|no|six|true|false|json|xml|yaml|csv|html", a, re.I) or len(a) <= 2:
        return {"storage": "cloud data storage", "networking": "computer network",
                "compute": "server computer", "identity": "login security",
                "security": "cyber security shield", "cost": "money budget",
                "governance": "settings management", "cloud": "cloud computing"}[cat]
    return a

# ---- generate ----
CHUNK = 40
os.makedirs("kahoot_images", exist_ok=True)
guide_rows = []  # (part, qno_in_part, overall, question, answer, cat_title, filename, keyword)
nfiles = (len(MC) + CHUNK - 1) // CHUNK
for part in range(nfiles):
    pdir = os.path.join("kahoot_images", f"Part{part+1}")
    os.makedirs(pdir, exist_ok=True)
    chunk = MC[part*CHUNK:(part+1)*CHUNK]
    for li, row in enumerate(chunk):
        q, ans = row[0], row[1]
        overall = part*CHUNK + li + 1
        cat = classify(q, ans)
        fname = f"Part{part+1}_Q{li+1:02d}.png"
        fpath = os.path.join(pdir, fname)
        footer = f"AZ-900 Review  -  Part {part+1}  -  Q{li+1}"
        make_image(fpath, cat, footer)
        guide_rows.append((part+1, li+1, overall, q, ans, CATS[cat][2],
                           f"kahoot_images/Part{part+1}/{fname}", keyword(cat, ans)))

# ---- Yes/No category images ----
yn_rows = []  # (qno, statement, answer_yesno, cat_title, filename, keyword)
yndir = os.path.join("kahoot_images", "YesNo")
os.makedirs(yndir, exist_ok=True)
for i, (stmt, ans, src) in enumerate(TF, start=1):
    cat = classify(stmt, "")
    fname = f"YesNo_Q{i:02d}.png"
    footer = f"AZ-900 Review  -  Yes/No  -  Q{i}"
    make_image(os.path.join(yndir, fname), cat, footer)
    yn_rows.append((i, stmt, "Yes" if ans == "True" else "No", CATS[cat][2],
                    f"kahoot_images/YesNo/{fname}", keyword(cat, "yes")))

# ---- guide workbook ----
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()
ws0 = wb.active
ws0.title = "READ ME"
ws0.column_dimensions["B"].width = 116
ws0["B1"] = "AZ-900 Kahoot - Per-Question Image Guide"
ws0["B1"].font = Font(bold=True, color="1F4E78", size=14)
for i, ln in enumerate([
    "",
    "Kahoot's spreadsheet importer CANNOT include images - they must be added in the editor after import.",
    "This guide makes that fast. Two ways to add an image to each question:",
    "",
    "OPTION A - Drag in the ready-made image (fastest):",
    "  In the kahoot editor, open a question, then drag the matching file from the 'kahoot_images/PartX' folder",
    "  into the question's media box. Files are named PartX_Qnn.png in the SAME order as the import file.",
    "",
    "OPTION B - Use Kahoot's built-in image library (real photos, free):",
    "  Click the media box > search Kahoot's image library using the 'Kahoot image search keyword' in this guide.",
    "",
    "Each generated image reflects the question's SUBJECT AREA (Storage, Networking, etc.) so it stays relevant",
    "without revealing the answer on screen. Images are generated (no third-party photos = no licensing issues).",
    "",
    "The 'Yes/No' sheet covers the Yes/No category (AZ-900_Kahoot_YesNo.xlsx); its images live in",
    "kahoot_images/YesNo and are named YesNo_Qnn.png in the same order as that file.",
], start=2):
    ws0.cell(row=i, column=2, value=ln)
ws0.sheet_view.showGridLines = False

for part in range(nfiles):
    ws = wb.create_sheet(f"Part {part+1}")
    headers = ["Q # (in part)", "Overall #", "Question", "Correct Answer",
               "Subject area", "Image file to drag in", "Kahoot image search keyword"]
    ws.append(headers)
    for c in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=c); cell.fill = HEAD_FILL; cell.font = HEAD_FONT
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = BORDER
    for r in [x for x in guide_rows if x[0] == part+1]:
        ws.append([r[1], r[2], r[3], r[4], r[5], r[6], r[7]])
    for i, w in enumerate([12, 10, 60, 34, 18, 34, 30], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for rr in range(2, ws.max_row+1):
        for cc in range(1, len(headers)+1):
            ws.cell(row=rr, column=cc).alignment = WRAP
            ws.cell(row=rr, column=cc).border = BORDER
    ws.freeze_panes = "A2"

# Yes/No category sheet
ws = wb.create_sheet("Yes-No")
headers = ["Q #", "Statement", "Answer (Yes/No)", "Subject area",
           "Image file to drag in", "Kahoot image search keyword"]
ws.append(headers)
for c in range(1, len(headers)+1):
    cell = ws.cell(row=1, column=c); cell.fill = HEAD_FILL; cell.font = HEAD_FONT
    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    cell.border = BORDER
for r in yn_rows:
    ws.append(list(r))
for i, w in enumerate([6, 70, 16, 18, 34, 30], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
for rr in range(2, ws.max_row+1):
    for cc in range(1, len(headers)+1):
        ws.cell(row=rr, column=cc).alignment = WRAP
        ws.cell(row=rr, column=cc).border = BORDER
ws.freeze_panes = "A2"

wb.save("AZ-900_Kahoot_Image_Guide.xlsx")

# report
from collections import Counter
catcount = Counter(classify(r[0], r[1]) for r in MC)
print("Generated MC images:", len(guide_rows))
print("Generated Yes/No images:", len(yn_rows))
print("Per part:", Counter(r[0] for r in guide_rows))
print("Subject-area distribution (MC):", dict(catcount))
print("Guide workbook: AZ-900_Kahoot_Image_Guide.xlsx")
