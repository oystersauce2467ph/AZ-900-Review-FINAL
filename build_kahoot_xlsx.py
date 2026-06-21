# -*- coding: utf-8 -*-
"""
Build Kahoot-IMPORTABLE .xlsx files for AZ-900 from the 3 reviewer PDFs.

WHY THIS VERSION:
  Kahoot's spreadsheet importer only accepts QUIZ (multiple-choice) questions.
  "Type Answer" questions cannot be imported (and are a paid feature), so the
  earlier 'Use Cases (Type Answer)' sheet was ignored by Kahoot.

  Fix: every use-case is converted into a multiple-choice question. The CORRECT
  option is the exact answer from the PDF; the WRONG options (distractors) are
  OTHER REAL Azure terms taken from the same PDFs - no invented facts.

IMPORTER RULES ENFORCED (per Kahoot support):
  - Question text  <= 95 characters
  - Each answer    <= 60 characters
  - 2-4 answers, >=1 correct, correct given as option number(s)
  - Time limit in {5,10,20,30,60,120}

OUTPUT (single-sheet files = unambiguous import; each well under the 200/kahoot cap):
  - AZ-900_Kahoot_MultipleChoice.xlsx   (all multiple-choice questions)
  - AZ-900_Kahoot_TrueFalse.xlsx        (all true/false questions)
  - AZ-900_Reference.xlsx               (full PDF answers + sources + redundancy)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MCTIME = 30
TFTIME = 20

def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER

def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# =====================================================================
# MULTIPLE-CHOICE DATA
# (question, correct, d1, d2, d3, source)   correct is option #1 here;
# options are shuffled deterministically per row when written.
# =====================================================================
MC = [
 # ---- Original MC that already existed in the PDFs (kept verbatim) ----
 ("Each of 10 departments needs a different payment option. Create what per department?",
  "A subscription", "A resource group", "A container instance", "A reservation", "P2 Q2"),
 ("What is used to grant permission to Azure Virtual Desktop resources?",
  "Role-based access control (RBAC) roles", "Tags", "Resource groups", "Application security groups", "P2 Q4"),
 ("You plan to create a virtual machine in Azure. Where will it be placed?",
  "In a resource group", "In a storage account", "In an administrative unit", "In an application group", "P2 Q6"),
 ("How many copies of data does an Azure Storage account using LRS keep?",
  "3", "4", "6", "9", "P2 Q10"),
 ("Which pairing counts as TRUE multi-factor authentication (MFA)?",
  "Password + fingerprint", "Fingerprint + retina scan", None, None, "P3 Practice Test Q14"),

 # ---- Converted from P2 use cases ----
 ("Which gives a private dedicated link to Azure that avoids the public internet?",
  "ExpressRoute", "VPN Gateway", "VNet peering", "Azure DNS", "P2 Q7"),
 ("Which service connects two Azure VNets with low-latency, high-bandwidth?",
  "Virtual network peering", "VPN Gateway", "ExpressRoute", "Azure Firewall", "P2 Q7"),
 ("Which service provides an encrypted tunnel to Azure over the public internet?",
  "VPN Gateway", "ExpressRoute", "VNet peering", "Azure DNS", "P2 Q7"),
 ("When extending on-premises to Azure, which resource defines the on-prem VPN device?",
  "Local Network Gateway", "Virtual Network Gateway", "VNet peering", "Azure Firewall", "P2 Q9"),
 ("Requiring a password AND a one-time passcode at sign-in is an example of what?",
  "Multi-Factor Authentication (MFA)", "Single sign-on (SSO)", "Conditional Access", "RBAC", "P2 Q11"),
 ("In IaaS, which component is the cloud provider's responsibility?",
  "Physical hosts, network, and datacenter", "Operating system", "Applications and data", "User accounts and identities", "P2 Q12"),
 ("Which service provides a full Windows desktop experience from the cloud?",
  "Azure Virtual Desktop", "Azure App Service", "Azure Functions", "VM Scale Sets", "P2 Q13"),
 ("____ ensures access to cloud resources in the event of a service failure.",
  "High availability", "Scalability", "Elasticity", "Agility", "P2 Q14"),
 ("In a SaaS model, what does the customer provide?",
  "Data and user access (accounts)", "The operating system", "Physical servers", "The virtualization layer", "P2 Q15"),
 ("In SaaS, which responsibility is shared between Microsoft and the customer?",
  "Identity and access management", "Physical datacenter", "Network infrastructure", "OS patching", "P2 Q16"),
 ("Migrating an on-prem server using lift-and-shift uses which cloud service model?",
  "IaaS (Infrastructure as a Service)", "PaaS", "SaaS", "FaaS", "P2 Q18"),
 ("Which cloud benefit lets you recover from a catastrophic failure?",
  "Disaster recovery", "Geo-distribution", "Scalability", "Elasticity", "P2 Q19"),
 ("Which cloud benefit deploys resources close to users globally?",
  "Geo-distribution", "Disaster recovery", "High availability", "Scalability", "P2 Q19"),
 ("Which cloud benefit adjusts resources to meet changing demand?",
  "Scalability", "Disaster recovery", "Geo-distribution", "Predictability", "P2 Q19"),
 ("____ enables Azure resources to be deployed close to users.",
  "Geo-distribution (Azure Regions)", "Disaster recovery", "High availability", "Fault tolerance", "P2 Q21"),
 ("When migrating a public website to Azure, what must you plan to do?",
  "Pay monthly usage costs (OpEx)", "Buy hardware upfront (CapEx)", "Purchase perpetual licenses", "Pay one fixed fee", "P2 Q22"),
 ("You must ensure resources can be created only in specific Azure regions. Use what?",
  "Azure Policy", "Azure Blueprints", "RBAC", "Resource locks", "P2 Q25"),
 ("Azure Resource Manager (ARM) templates use which file format?",
  "JSON", "YAML", "XML", "HTML", "P2 Q28"),
 ("You can assign ____ to every Azure resource.",
  "Tags", "Locks", "Policies", "Roles", "P2 Q31"),
 ("What can identify underutilized or unused Azure virtual machines?",
  "Azure Advisor", "Azure Monitor", "Azure Service Health", "Azure Policy", "P2 Q32"),
 ("Which tool do you use to monitor the health of Azure services?",
  "Azure Service Health", "Azure Advisor", "Azure Monitor", "Microsoft Defender for Cloud", "P2 Q33"),
 ("Which tool shows security recommendations?",
  "Microsoft Defender for Cloud", "Azure Advisor", "Azure Monitor", "Azure Service Health", "P2 Q33"),
 ("You must prevent accidental deletion of resources. Which setting should you use?",
  "Delete lock (CanNotDelete)", "Read-only lock", "A tag", "Azure Policy", "P2 Q34"),
 ("____ extends Azure compliance and monitoring to hybrid and multicloud.",
  "Azure Arc", "Azure Migrate", "Azure Monitor", "Azure VMware Solution", "P2 Q35"),
 ("Correct order of steps to deploy 5 identical VMs from an existing VM?",
  "Generalize, Image, ARM template, Deploy", "Deploy, Image, Generalize, Template",
  "Image, Deploy, Generalize, Template", "Template, Deploy, Image, Generalize", "P2 Q36"),

 # ---- Converted from P3 domain & practice questions ----
 ("Which cloud model is considered the most flexible?",
  "Hybrid", "Public", "Private", "Community", "P3 D1 Q1"),
 ("Which expenditure model means paying for resources only as they are used?",
  "OpEx (consumption-based)", "CapEx", "Reserved pricing", "Perpetual licensing", "P3 D1 Q2"),
 ("What term describes the automatic addition of resources based on demand?",
  "Elasticity", "Scalability", "Agility", "Reliability", "P3 D1 Q3"),
 ("Microsoft 365 Online falls under which cloud service category?",
  "SaaS", "IaaS", "PaaS", "FaaS", "P3 D1 Q4"),
 ("Which cloud model gives total control over security and resources?",
  "Private", "Public", "Hybrid", "Community", "P3 D1 Q5"),
 ("What is the minimum number of Availability Zones in a supporting region?",
  "3", "1", "2", "5", "P3 D2 Q1"),
 ("Which storage tier is for data stored at least 180 days and rarely accessed?",
  "Archive", "Hot", "Cool", "Cold", "P3 D2 Q2"),
 ("Which Azure service is a Layer 7 load balancer for web applications?",
  "Azure Application Gateway", "Azure Load Balancer", "Azure Firewall", "Azure DNS", "P3 D2 Q3"),
 ("If a resource group is deleted, what happens to the resources inside it?",
  "They are all deleted", "They move to a default group", "They become read-only", "Nothing happens", "P3 D2 Q4"),
 ("What physical tool transfers massive data (e.g., 40 TB) to Azure?",
  "Azure Data Box", "AzCopy", "Azure File Sync", "Azure Migrate", "P3 D2 Q5"),
 ("Which pricing model gives the biggest discount for VMs reclaimable anytime?",
  "Spot Instances", "Reserved Instances", "Azure Hybrid Benefit", "Pay-as-you-go", "P3 D3 Q1"),
 ("Which website provides Microsoft's audit reports and compliance info?",
  "Service Trust Portal", "Trust Center", "Azure Advisor", "Microsoft Purview", "P3 D3 Q2"),
 ("What is the name of a collection of Azure policies grouped together?",
  "An Initiative", "A Blueprint", "A Resource Group", "A Management Group", "P3 D3 Q3"),
 ("Under Shared Responsibility, what is the provider ALWAYS responsible for?",
  "Physical datacenters and network", "Data and accounts", "Access management", "Operating system", "P3 D3 Q4"),
 ("Which tool recommends improvements for cost, security, reliability, performance?",
  "Azure Advisor", "Azure Monitor", "Azure Service Health", "Azure Policy", "P3 D3 Q5"),
 ("AD stays on-prem but the database is hosted in the cloud. Which cloud model?",
  "Hybrid cloud", "Public cloud", "Private cloud", "Community cloud", "P3 PT Q1"),
 ("Which cloud service type gives a dev environment without managing the OS?",
  "PaaS", "IaaS", "SaaS", "FaaS", "P3 PT Q2"),
 ("Which networking service reaches Microsoft datacenters without public internet?",
  "ExpressRoute", "VPN Gateway", "VNet peering", "Azure DNS", "P3 PT Q3"),
 ("An admin must restrict which regions resources deploy into. What to set?",
  "Azure Policy", "RBAC", "A resource lock", "A Blueprint", "P3 PT Q4"),
 ("Which redundancy offers 16 nines by copying to a secondary region?",
  "Geo-Redundant Storage (GRS)", "Locally Redundant (LRS)", "Zone Redundant (ZRS)", "Standard GPv2", "P3 PT Q5"),
 ("A user signs in to Azure with a username and password. Which process occurred?",
  "Authentication", "Authorization", "Conditional Access", "Encryption", "P3 PT Q6"),
 ("When a resource is moved to a different resource group, its permissions...?",
  "Inherit the new group's permissions", "Stay the same", "Are removed", "Become read-only", "P3 PT Q7"),
 ("What file format builds Azure Resource Manager (ARM) templates?",
  "JSON", "XML", "YAML", "CSV", "P3 PT Q8"),
 ("Cheapest VM for an experiment that can be shut down anytime?",
  "Spot Instances", "Reserved Instances", "Azure Hybrid Benefit", "A dedicated host", "P3 PT Q9"),
 ("Which two factors does Azure use to price storage disks?",
  "Quality (SSD/HDD) and size", "Region and OS", "CPU and RAM", "Bandwidth and IOPS", "P3 PT Q10"),
 ("A business wants total control of its hardware for a single-org app. Which model?",
  "Private cloud", "Public cloud", "Hybrid cloud", "Community cloud", "P3 PT Q11"),
 ("Which tool should you check FIRST for a suspected Azure-wide outage?",
  "Azure Service Health", "Azure Advisor", "Azure Monitor", "Azure Policy", "P3 PT Q12"),
 ("A school must isolate traffic for each classroom within one VNet. Create what?",
  "Virtual subnets", "VNet peering", "Network Security Groups", "VPN gateways", "P3 PT Q13"),
 ("Minimum subscription tier required to use Microsoft Defender for Cloud?",
  "Free", "Standard", "Professional Direct", "Premier", "P3 PT Q15"),

 # ---- Converted from P1 fundamentals notes ----
 ("What is the basic, deployable building block of Azure?",
  "A resource", "A subscription", "A management group", "A region", "P1 p2"),
 ("Every resource must belong to how many resource groups?",
  "Exactly one", "Zero", "Two", "Unlimited", "P1 p2"),
 ("Can resource groups be nested inside one another?",
  "No", "Yes", "Only inside management groups", "Only with a lock", "P1 p2"),
 ("Can a resource group be renamed after it is created?",
  "No", "Yes", "Only via the CLI", "Only when empty", "P1 p2"),
 ("How many management groups can a single directory support?",
  "10,000", "100", "1,000", "Unlimited", "P1 p2"),
 ("Management groups support up to how many levels of depth?",
  "Six", "Three", "Ten", "Unlimited", "P1 p2"),
 ("What are the two subscription boundary types?",
  "Billing and access control", "Region and zone", "Compute and storage", "Identity and network", "P1 p2"),
 ("Which service is a container orchestration service for many containers?",
  "Azure Kubernetes Service (AKS)", "Azure Functions", "Azure App Service", "VM Scale Sets", "P1 p3"),
 ("Which service runs event-driven serverless code only when triggered?",
  "Azure Functions", "Azure App Service", "AKS", "Azure Batch", "P1 p3"),
 ("Azure Functions runs in response to which triggers?",
  "HTTP, timers, and service messages", "Only manual runs", "Only VM reboots", "Only disk writes", "P1 p3"),
 ("What does a VM Scale Set provide?",
  "A group of identical load-balanced VMs", "A single very large VM", "Serverless functions", "Container orchestration", "P1 p2"),
 ("VM availability sets group virtual machines by what?",
  "Update and fault domains", "Regions and zones", "Subnets and VNets", "Tags and locks", "P1 p2"),
 ("What is an update domain?",
  "VMs rebooted together for maintenance", "VMs sharing a power source", "VMs in different regions", "VMs with the same tag", "P1 p2"),
 ("What is a fault domain?",
  "VMs sharing a power/network failure point", "VMs rebooted together", "VMs in separate regions", "VMs in one subnet", "P1 p2"),
 ("Which IoT service enables secure two-way cloud-to-device communication?",
  "Azure IoT Hub", "Azure IoT Central", "Azure IoT Edge", "Azure Arc", "P1 p3"),
 ("Which IoT service is a SaaS IoT platform for solution builders?",
  "Azure IoT Central", "Azure IoT Hub", "Azure IoT Edge", "Azure Sphere", "P1 p3"),
 ("Which IoT service extends cloud capabilities to edge devices?",
  "Azure IoT Edge", "Azure IoT Hub", "Azure IoT Central", "Azure Arc", "P1 p3"),
 ("Which Azure service supports generative AI like chat and content generation?",
  "Azure OpenAI Service", "Azure Machine Learning", "Azure Cognitive Search", "Azure Bot Service", "P1 p3"),
 ("What are Network Security Groups (NSGs) used for?",
  "Allow/block traffic via in/outbound rules", "Encrypt data at rest", "Store secrets", "Balance web traffic", "P1 p3"),
 ("What is a network virtual appliance?",
  "A VM that performs a network function", "A physical router", "A storage account", "A DNS record", "P1 p3"),
 ("Which ExpressRoute type connects to the Microsoft backbone at 100 Gbps?",
  "ExpressRoute Direct", "VPN Gateway", "Site-to-Site VPN", "Point-to-Site VPN", "P1 p4"),
 ("What does the Azure VPN Gateway use to connect networks?",
  "The public internet (encrypted)", "A private Microsoft link", "A physical cable", "A satellite link", "P1 p4"),
 ("Point-to-Site vs Site-to-Site: which connects an entire network?",
  "Site-to-Site", "Point-to-Site", "ExpressRoute Direct", "VNet peering", "P1 p3"),
 ("What is the DEFAULT VPN gateway high-availability option?",
  "Active/Standby", "Active/Active", "Load-balanced", "Geo-redundant", "P1 p4"),
 ("What are alias records in Azure DNS?",
  "Records that auto-update with Azure resources", "Static IP mappings", "Encrypted secrets", "Firewall rules", "P1 p4"),
 ("Azure storage account names must follow which rule?",
  "3-24 lowercase letters/numbers, unique", "Any case, up to 50 chars", "Uppercase letters only", "Spaces allowed", "P1 p4"),
 ("Which storage account type is the default for most scenarios?",
  "Standard general-purpose v2", "Premium block blobs", "Premium file shares", "Premium page blobs", "P1 p4"),
 ("LRS keeps how many copies and where?",
  "3 copies in one datacenter", "3 across availability zones", "6 across regions", "1 copy locally", "P1 p4"),
 ("ZRS keeps how many copies and where?",
  "3 copies across availability zones", "3 in one datacenter", "6 across regions", "1 copy locally", "P1 p4"),
 ("GRS keeps copies where?",
  "3 local plus 3 in a paired region", "3 in one datacenter only", "Across zones only", "1 copy locally", "P1 p4"),
 ("What does RA-GRS add over GRS?",
  "Read access to the secondary region", "More local copies", "Lower cost", "Zone redundancy", "P1 p4"),
 ("Which storage service holds unstructured data like media and binary files?",
  "Azure Blob", "Azure Tables", "Azure Queues", "Azure Files", "P1 p4"),
 ("Which storage service is NoSQL for structured, non-relational data?",
  "Azure Tables", "Azure Blob", "Azure Queues", "Azure Disks", "P1 p4"),
 ("Which storage service provides asynchronous messaging between apps?",
  "Azure Queues", "Azure Blob", "Azure Tables", "Azure Files", "P1 p4"),
 ("Which storage service provides block-level virtual disks for VMs?",
  "Azure Disks", "Azure Blob", "Azure Files", "Azure Queues", "P1 p4"),
 ("Which storage service manages cloud file shares using SMB and NFS?",
  "Azure Files", "Azure Blob", "Azure Tables", "Azure Disks", "P1 p4"),
 ("The Hot blob tier is best for what access pattern?",
  "Frequent, daily access", "Infrequent (30+ days)", "Rare (90+ days)", "Long-term (180+ days)", "P1 p4"),
 ("The Cool blob tier minimum storage duration is?",
  "30 days", "7 days", "90 days", "180 days", "P1 p4"),
 ("The Cold blob tier minimum storage duration is?",
  "90 days", "30 days", "180 days", "7 days", "P1 p4"),
 ("The Archive blob tier minimum storage duration is?",
  "180 days", "30 days", "90 days", "365 days", "P1 p4"),
 ("What is the maximum usable capacity of an Azure Data Box device?",
  "80 terabytes", "40 terabytes", "8 terabytes", "1 petabyte", "P1 p5"),
 ("Which standard is used to wipe Data Box disks after upload?",
  "NIST 800-88r1", "ISO 27018", "GDPR", "FedRAMP", "P1 p5"),
 ("What is AzCopy?",
  "Command-line tool to copy blobs/files", "A GUI storage app", "A backup service", "A VM image tool", "P1 p5"),
 ("What is Azure Storage Explorer?",
  "GUI app to manage files and blobs", "A command-line tool", "A monitoring service", "A firewall", "P1 p5"),
 ("What two features does Azure File Sync provide?",
  "Bidirectional sync and cloud tiering", "Encryption and backup", "Load balancing and DNS", "Logging and alerts", "P1 p5"),
 ("What is Microsoft Entra ID?",
  "Cloud identity and access management", "A storage service", "A VM type", "A monitoring tool", "P1 p5"),
 ("What bridges on-prem Active Directory with Microsoft Entra ID?",
  "Microsoft Entra Connect", "Entra Domain Services", "Azure Arc", "Azure Migrate", "P1 p5"),
 ("An Entra Domain Services managed domain syncs in which direction?",
  "One-way, from Entra ID", "Two-way", "On-prem to cloud only", "No sync", "P1 p5"),
 ("Which is one of the four Azure authentication methods?",
  "Passwordless", "Encryption", "Tagging", "Load balancing", "P1 p6"),
 ("Which is an Azure passwordless sign-in option?",
  "Windows Hello", "Password plus PIN", "A security question", "An email link", "P1 p6"),
 ("MFA factors are based on something you know, something you are, and...?",
  "Something you own", "Something you see", "Somewhere you live", "Someone you know", "P1 p6"),
 ("Microsoft Defender for Cloud helps you assess, secure, and...?",
  "Defend", "Delete", "Deploy", "Discount", "P1 p6"),
 ("Which is a layer of the Defense in Depth model?",
  "Perimeter", "Subscription", "Resource group", "Initiative", "P1 p6"),
 ("What is Azure Policy?",
  "Enforces and audits resource compliance", "Assigns user roles", "Stores secrets", "Balances traffic", "P1 p6"),
 ("Which is a valid Azure Policy effect?",
  "Deny", "Encrypt", "Scale", "Backup", "P1 p6"),
 ("What is an Azure Initiative?",
  "A collection of Azure policies", "A group of subscriptions", "A backup plan", "A VM image", "P1 p7"),
 ("What is an Azure Blueprint?",
  "A set of standards/templates to deploy", "A single policy", "A storage tier", "A VM size", "P1 p7"),
 ("What does a resource tag consist of?",
  "A name and a value", "A key only", "A role and scope", "A lock and level", "P1 p7"),
 ("What does a DoNotDelete lock allow?",
  "Read and modify, but not delete", "Read only", "Delete only", "Nothing at all", "P1 p7"),
 ("What does a ReadOnly lock allow?",
  "Read only, no modify or delete", "Modify but not delete", "Delete only", "Full access", "P1 p7"),
 ("What syntax style does the Azure CLI use?",
  "Bash-style syntax", "PowerShell cmdlets", "JSON only", "C# code", "P1 p7"),
 ("What does Azure PowerShell use?",
  "Cmdlets and REST APIs", "Bash scripts", "SQL queries", "YAML files", "P1 p7"),
 ("What is Azure Resource Manager (ARM)?",
  "The service all resource actions go through", "A storage account", "A monitoring tool", "A VM type", "P1 p7"),
 ("Which two declarative languages are used for ARM templates?",
  "JSON and Bicep", "YAML and XML", "Bash and PowerShell", "C# and Python", "P1 p7"),
 ("What is Azure Key Vault used for?",
  "Storing secrets, keys, and certificates", "Monitoring metrics", "Balancing traffic", "Running VMs", "P1 p7"),
 ("Which is an Azure Advisor recommendation category?",
  "Cost", "Tagging", "Naming", "Billing currency", "P1 p7"),
 ("Which is a component of Azure Service Health?",
  "Resource Health", "Log Analytics", "Application Insights", "Cloud Shell", "P1 p7"),
 ("What is Azure Monitor?",
  "Collects telemetry/metrics and alerts", "Stores secrets", "Assigns roles", "Hosts web apps", "P1 p8"),
 ("What is Azure Log Analytics?",
  "Tool to query data Monitor collects", "A storage tier", "A firewall", "A VM image", "P1 p8"),
 ("What is Application Insights?",
  "Monitors web app performance and usage", "Stores blobs", "Manages identities", "Enforces policy", "P1 p8"),
 ("Which factor affects Azure cost?",
  "Resource type, usage, and geography", "The user's browser", "Time of day only", "Number of admins", "P1 p6"),
 ("What is the Azure Pricing Calculator used for?",
  "Estimate cost of future Azure services", "Compare to on-premises", "Track actual spend", "Set budgets", "P1 p6"),
 ("What is the TCO Calculator used for?",
  "Compare on-premises vs Azure costs", "Estimate a monthly bill", "Track actual spend", "Set alerts", "P1 p6"),
 ("Which is a type of Azure cost alert?",
  "Budget alert", "Latency alert", "Login alert", "Region alert", "P1 p6"),
 ("What is Microsoft Purview?",
  "Unified data governance and compliance", "A VM type", "A storage tier", "A load balancer", "P1 p6"),
 ("Azure Reservations are best for what kind of workloads?",
  "Stable, predictable workloads", "Interruptible test VMs", "One-time jobs", "Bursty traffic", "P1 p8"),
 ("What is the Azure savings plan for compute?",
  "Commit to hourly spend for 1-3 years", "Pay-per-second billing", "Free tier credits", "Spot bidding", "P1 p8"),
 ("What are Spot Virtual Machines?",
  "Unused capacity at reduced, interruptible prices", "Reserved 3-year VMs", "Dedicated hosts", "Always-on premium VMs", "P1 p8"),
 ("What is Azure Cloud Shell?",
  "Browser shell for PowerShell and CLI", "A desktop app", "A VM image", "A storage tool", "P1 p9"),
 ("What is Azure Migrate?",
  "A unified on-prem to cloud migration tool", "A backup service", "A monitoring tool", "A DNS service", "P1 p5"),
 ("What is Azure Arc?",
  "Manages on-prem and multicloud resources", "Hosts web apps", "Stores blobs", "Balances traffic", "P1 p1"),
 ("What is Azure VMware Solution?",
  "Runs VMware workloads on Azure", "Runs Linux containers", "A storage tier", "A DNS service", "P1 p1"),
 ("What is the Service Trust Portal?",
  "Access to security/privacy/compliance docs", "A billing dashboard", "A VM console", "A code editor", "P1 p7"),
 ("How does the reviewer define cloud computing?",
  "Compute/network/storage over the Internet", "Buying servers for a datacenter", "Running apps only offline", "A type of VPN", "P1 p1"),
 ("Which is a listed IaaS use case?",
  "Lift-and-shift migration", "Messaging", "Productivity apps", "Expense tracking", "P1 p1"),
 ("Which is a listed PaaS use case?",
  "Development framework", "Lift-and-shift migration", "Messaging", "Finance tracking", "P1 p1"),
 ("Which is a listed SaaS use case?",
  "Messaging and productivity", "Lift-and-shift migration", "Development framework", "Testing/deployment", "P1 p1"),
 ("Which is ALWAYS the company's responsibility?",
  "Accounts and identities", "Physical hosts", "Physical network", "The datacenter", "P1 p1"),
 ("Which is ALWAYS the cloud provider's responsibility?",
  "The physical datacenter", "Your data", "Accounts and identities", "Your devices", "P1 p1"),
 ("What is Capital Expenditure (CapEx)?",
  "One-time upfront hardware costs", "Pay-as-you-go spending", "Monthly usage fees", "Per-second billing", "P1 p1"),
 ("What is Operational Expenditure (OpEx)?",
  "Pay-as-you-go spending over time", "Upfront hardware costs", "A perpetual license", "A one-time fee", "P1 p1"),
 ("How is an Azure SLA expressed?",
  "As a percentage of uptime", "In dollars", "In data copies", "In regions", "P1 p1"),
 ("Per the reviewer, 99% uptime means how much downtime per month?",
  "7.2 hours per month", "45 minutes", "4 hours", "24 hours", "P1 p1"),
 ("Which benefit means resources stay accessible during failures and updates?",
  "High Availability", "Elasticity", "Agility", "Predictability", "P1 p1"),
 ("Which benefit increases/decreases capacity based on workload?",
  "Scalability", "Reliability", "Security", "Compliance", "P1 p1"),
 ("What is vertical scaling (scale up/down)?",
  "Changing a resource's CPU/RAM", "Adding more instances", "Removing a region", "Adding a subnet", "P1 p1"),
 ("What is horizontal scaling (scale in/out)?",
  "Adding or removing instances", "Increasing VM CPU/RAM", "Changing the region", "Adding storage tiers", "P1 p1"),
 ("Which benefit is automatic scalability?",
  "Elasticity", "Agility", "Reliability", "Predictability", "P1 p1"),
 ("Which benefit means recovering after system failures?",
  "Reliability", "Elasticity", "Agility", "Sustainability", "P1 p1"),
 ("Which benefit covers forecasting performance or costs?",
  "Predictability", "Elasticity", "Agility", "Reliability", "P1 p1"),
 ("Which benefit covers encryption, identity, and threat monitoring?",
  "Security", "Scalability", "Agility", "Predictability", "P1 p1"),
 ("Which benefit involves auditing and flagging non-compliant resources?",
  "Compliance", "Elasticity", "Agility", "Reliability", "P1 p1"),
 ("Which benefit uses templates for corporate/government standards?",
  "Governance", "Elasticity", "Scalability", "Agility", "P1 p1"),
 ("Which benefit covers deploying resources, tracking, and alerts?",
  "Manageability", "Elasticity", "Reliability", "Security", "P1 p1"),
 ("Which benefit reduces carbon footprint via economies of scale?",
  "Sustainability", "Elasticity", "Agility", "Predictability", "P1 p1"),
 ("What are Azure regions?",
  "Geographic areas containing datacenters", "Single servers", "Storage tiers", "VM sizes", "P1 p2"),
 ("Which service is NOT region-locked?",
  "Microsoft Entra ID", "Azure Virtual Machines", "Azure Disks", "A SQL Server on a VM", "P1 p2"),
 ("What are availability zones?",
  "Physically separate datacenters in a region", "Two regions paired together", "Subnets in a VNet", "Copies of a blob", "P1 p2"),
 ("Which is a service category regarding availability zones?",
  "Zone-redundant", "Region-locked", "Tag-based", "Policy-based", "P1 p2"),
 ("What is the separation distance for region pairs?",
  "About 300 miles (482 km)", "About 10 miles", "About 1,000 miles", "The same building", "P1 p2"),
 ("What are sovereign regions used for?",
  "Isolation for compliance/legal reasons", "Cheaper pricing", "Faster networking", "Extra storage", "P1 p2"),
 ("What does Azure Load Balancer do?",
  "Distributes web traffic across VMs", "Stores secrets", "Assigns roles", "Encrypts disks", "P1 p2"),
 ("What are the four Azure App Service app types?",
  "Web, API, WebJobs, Mobile apps", "VMs, disks, blobs, queues", "Hot, Cool, Cold, Archive", "LRS, ZRS, GRS, GZRS", "P1 p3"),

 # ---- P3 extra concepts ----
 ("What does the Zero Trust model assume?",
  "A breach is happening; verify by auth", "The network is safe", "Location grants trust", "No auth is needed", "P3 3.1"),
 ("Conditional Access grants or denies access based on which signals?",
  "Identity, device, and location", "CPU and RAM", "Region and zone", "Cost and budget", "P3 3.2"),
 ("What are the two Azure DDoS Protection tiers?",
  "Basic (free) and Standard (paid)", "Bronze and Gold", "Hot and Cool", "Free and Premier", "P3 3.5"),
 ("Which support plan offers 24x7 phone and email support?",
  "Standard", "Basic/Developer", "Professional Direct", "Premier", "P3 3.9"),
 ("Which support plan includes support for some non-Microsoft products?",
  "Premier", "Basic", "Developer", "Standard", "P3 3.9"),
 ("Which is a subscription type covered by the reviewer?",
  "Enterprise Agreement", "Spot subscription", "Region pair", "Resource lock", "P3 3.9"),
 ("What is RPO (Recovery Point Objective)?",
  "Maximum acceptable data loss, in time", "Time to restore service", "Uptime percentage", "Number of backups", "P3 3.9"),
 ("What is RTO (Recovery Time Objective)?",
  "How long it takes to restore service", "Maximum acceptable data loss", "Uptime percentage", "Backup frequency", "P3 3.9"),
 ("Correct order of the Azure feature lifecycle stages?",
  "Private Preview, Public Preview, GA", "GA, Public Preview, Private Preview", "Public Preview, GA, Private", "Beta, Alpha, GA", "P3 3.9"),
 ("What is the Azure Hybrid Benefit?",
  "Reuse on-prem licenses in Azure", "Free 12 months of services", "Spot VM discounts", "A 3-year reservation", "P3 3.8"),
 ("What protects against a single datacenter's reboot or hardware fault?",
  "An Availability Set", "An Availability Zone", "A Region Pair", "A subscription", "P3 2.1"),
]

# =====================================================================
# TRUE / FALSE DATA  (statement, "True"/"False", source)
# =====================================================================
TF = [
 ("You can create custom Azure roles to control access to resources.", "True", "P2 Q1"),
 ("A user account can be assigned to multiple Azure roles.", "True", "P2 Q1"),
 ("A resource group can have the Owner role assigned to multiple users.", "True", "P2 Q1"),
 ("An Azure subscription can have multiple account administrators.", "True", "P2 Q3"),
 ("An Azure subscription can be managed by using a Microsoft account only.", "False", "P2 Q3"),
 ("An Azure resource group can contain multiple Azure subscriptions.", "False", "P2 Q3"),
 ("Single sign-on requires all users to sign in with Microsoft Authenticator.", "False", "P2 Q5"),
 ("Authentication establishes which level of access an authenticated user has.", "False", "P2 Q5"),
 ("Conditional Access uses sign-in signals to allow or deny access.", "True", "P2 Q5"),
 ("Data in an Azure Storage account automatically has at least three copies.", "True", "P2 Q8"),
 ("All data in an Azure Storage account is auto-backed up to another datacenter.", "False", "P2 Q8"),
 ("An Azure Storage account can hold only up to 2 TB and one million files.", "False", "P2 Q8"),
 ("Azure Reservations cost less than pay-as-you-go pricing.", "True", "P2 Q17"),
 ("Two Azure VMs of the same size always have the same monthly cost.", "False", "P2 Q17"),
 ("When an Azure virtual machine is stopped, storage costs still apply.", "True", "P2 Q17"),
 ("Azure Files is an example of Infrastructure as a Service (IaaS).", "True", "P2 Q20"),
 ("A DNS server on an Azure VM is an example of Platform as a Service (PaaS).", "False", "P2 Q20"),
 ("Microsoft Intune is an example of Software as a Service (SaaS).", "True", "P2 Q20"),
 ("Creating and configuring a virtual network is part of the PaaS model.", "False", "P2 Q23"),
 ("Updating application code in Azure App Service is the customer's job.", "True", "P2 Q23"),
 ("Configuring user access in PaaS is the customer's responsibility.", "True", "P2 Q23"),
 ("A resource lock can be applied to a Microsoft Entra user.", "False", "P2 Q24"),
 ("Multiple resource locks can be applied to the same virtual machine.", "True", "P2 Q24"),
 ("A delete lock allows modifying a resource but prevents deletion.", "True", "P2 Q24"),
 ("Azure PowerShell can be installed on macOS.", "True", "P2 Q26"),
 ("Azure Cloud Shell can be accessed from a Linux computer using a browser.", "True", "P2 Q26"),
 ("The Azure portal can be accessed only from Windows devices.", "False", "P2 Q26"),
 ("Inbound traffic to Azure using ExpressRoute is always free.", "True", "P2 Q27"),
 ("Outbound traffic from Azure to on-premises networks is always free.", "False", "P2 Q27"),
 ("Data transfer between Azure services in the same region is always free.", "True", "P2 Q27"),
 ("Azure Monitor can monitor the performance of on-premises computers.", "True", "P2 Q29"),
 ("Azure Monitor can send alerts to Microsoft Entra security groups.", "True", "P2 Q29"),
 ("Azure Monitor can trigger alerts based on Log Analytics data.", "True", "P2 Q29"),
 ("Only one tag can be assigned to an Azure resource.", "False", "P2 Q30"),
 ("Tags can be assigned to resources using ARM templates.", "True", "P2 Q30"),
 ("Tags can be used to enforce naming standards.", "False", "P2 Q30"),
]

# =====================================================================
# Build MC importable file (deterministic option shuffle so correct
# answer is not always option 1)
# =====================================================================
def build_mc_file():
    wb = Workbook()
    ws = wb.active
    ws.title = "Kahoot Quiz Import"
    headers = ["Question", "Answer 1", "Answer 2", "Answer 3", "Answer 4",
               "Time limit (sec)", "Correct answer(s)"]
    ws.append(headers)
    style_header(ws, len(headers))
    errors = []
    for idx, row in enumerate(MC):
        q, correct, *rest = row[0], row[1], *([row[2], row[3], row[4]])
        source = row[5]
        opts = [correct] + [o for o in rest if o is not None]
        # deterministic rotation so the correct answer position varies
        shift = idx % len(opts)
        rotated = opts[shift:] + opts[:shift]
        correct_pos = rotated.index(correct) + 1
        # pad to 4 columns
        padded = rotated + [None] * (4 - len(rotated))
        ws.append([q] + padded[:4] + [MCTIME, str(correct_pos)])
        # validation
        if len(q) > 95:
            errors.append(f"Q>{95} ({len(q)}): {q}")
        for o in rotated:
            if o and len(o) > 60:
                errors.append(f"ANS>{60} ({len(o)}): {o}")
    widths(ws, [70, 30, 30, 30, 30, 14, 16])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).alignment = WRAP
            ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=6).alignment = CENTER
        ws.cell(row=r, column=7).alignment = CENTER
    ws.freeze_panes = "A2"
    wb.save("AZ-900_Kahoot_MultipleChoice.xlsx")
    return errors

def build_tf_file():
    wb = Workbook()
    ws = wb.active
    ws.title = "Kahoot Quiz Import"
    headers = ["Question", "Answer 1", "Answer 2", "Answer 3", "Answer 4",
               "Time limit (sec)", "Correct answer(s)"]
    ws.append(headers)
    style_header(ws, len(headers))
    errors = []
    for stmt, ans, src in TF:
        correct_pos = "1" if ans == "True" else "2"
        ws.append([stmt, "True", "False", None, None, TFTIME, correct_pos])
        if len(stmt) > 95:
            errors.append(f"TF Q>{95} ({len(stmt)}): {stmt}")
    widths(ws, [80, 12, 12, 10, 10, 14, 16])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).alignment = WRAP
            ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=6).alignment = CENTER
        ws.cell(row=r, column=7).alignment = CENTER
    ws.freeze_panes = "A2"
    wb.save("AZ-900_Kahoot_TrueFalse.xlsx")
    return errors

def build_reference_file():
    wb = Workbook()
    # READ ME
    ws = wb.active
    ws.title = "READ ME"
    widths(ws, [4, 118])
    ws["B1"] = "AZ-900 Kahoot - Reference & Import Guide"
    ws["B1"].font = Font(bold=True, color="1F4E78", size=14)
    lines = [
        "",
        "WHY THERE ARE TWO IMPORT FILES:",
        "  Kahoot's spreadsheet importer accepts ONLY multiple-choice (quiz) questions, and the free plan",
        "  cannot import 'Type Answer' questions at all. So every 'use case' was converted into a",
        "  multiple-choice question: the CORRECT option is the exact answer from the PDF, and the wrong",
        "  options are OTHER REAL Azure terms taken from the same PDFs (no invented facts).",
        "",
        "IMPORT THESE TWO FILES (each is one clean sheet in Kahoot's template format):",
        "  1. AZ-900_Kahoot_MultipleChoice.xlsx  -> import as one kahoot (multiple-choice questions).",
        "  2. AZ-900_Kahoot_TrueFalse.xlsx       -> import as a second kahoot (True/False questions).",
        "  Keeping them in two files keeps each kahoot under Kahoot's 200-question limit.",
        "",
        "HOW TO IMPORT:",
        "  Create a new kahoot > Add question > 'Import questions from spreadsheet' > upload the file.",
        "  Correct answer column uses option NUMBERS (1-4). For True/False: 1 = True, 2 = False.",
        "",
        "THIS REFERENCE WORKBOOK CONTAINS:",
        "  - 'MC (with source + answer)'  : every multiple-choice question, its correct answer, and the PDF source.",
        "  - 'True-False (with source)'   : every True/False statement, the answer, and the PDF source.",
        "  - 'Original Use-Case Answers'  : the open-ended answers exactly as stated in the PDFs (for study).",
        "  - 'Redundant Questions'        : duplicate concepts to avoid repeating.",
        "  - 'Source Map'                 : which PDF each block came from.",
        "",
        "INTEGRITY: every question and answer comes strictly from the 3 PDFs. Wrong MC options are real",
        "Azure terms drawn from those same PDFs, used only as distractors - no facts were invented.",
    ]
    r = 2
    for ln in lines:
        ws.cell(row=r, column=2, value=ln)
        if ln.endswith(":"):
            ws.cell(row=r, column=2).font = Font(bold=True, color="1F4E78")
        r += 1
    ws.sheet_view.showGridLines = False

    # MC reference
    ws = wb.create_sheet("MC (with source + answer)")
    hdr = ["#", "Question", "Correct Answer (from PDF)", "Other options (real Azure terms)", "Source (PDF)"]
    ws.append(hdr); style_header(ws, len(hdr))
    for i, row in enumerate(MC, start=1):
        q, correct = row[0], row[1]
        others = [o for o in (row[2], row[3], row[4]) if o is not None]
        ws.append([i, q, correct, "; ".join(others), row[5]])
    widths(ws, [5, 66, 40, 48, 16])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(hdr) + 1):
            ws.cell(row=r, column=c).alignment = WRAP
            ws.cell(row=r, column=c).border = BORDER

    # TF reference
    ws = wb.create_sheet("True-False (with source)")
    hdr = ["#", "Statement", "Answer", "Source (PDF)"]
    ws.append(hdr); style_header(ws, len(hdr))
    for i, (stmt, ans, src) in enumerate(TF, start=1):
        ws.append([i, stmt, ans, src])
    widths(ws, [5, 80, 10, 14])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(hdr) + 1):
            ws.cell(row=r, column=c).alignment = WRAP
            ws.cell(row=r, column=c).border = BORDER

    # Original use-case answers (full, for study / in case user upgrades to Type Answer)
    ws = wb.create_sheet("Original Use-Case Answers")
    hdr = ["#", "Topic", "Question", "Full Answer (verbatim from PDF)", "Source"]
    ws.append(hdr); style_header(ws, len(hdr))
    for i, (topic, q, ans, src) in enumerate(USECASES, start=1):
        ws.append([i, topic, q, ans, src])
    widths(ws, [5, 18, 64, 52, 16])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(hdr) + 1):
            ws.cell(row=r, column=c).alignment = WRAP
            ws.cell(row=r, column=c).border = BORDER

    # Redundant
    ws = wb.create_sheet("Redundant Questions")
    hdr = ["Concept (asked more than once)", "Where it appears", "Recommendation"]
    ws.append(hdr); style_header(ws, len(hdr))
    for row in REDUNDANT:
        ws.append(list(row))
    widths(ws, [50, 44, 52])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(hdr) + 1):
            ws.cell(row=r, column=c).alignment = WRAP
            ws.cell(row=r, column=c).border = BORDER

    # Source map
    ws = wb.create_sheet("Source Map")
    hdr = ["Code", "PDF File", "What it contributed"]
    ws.append(hdr); style_header(ws, len(hdr))
    for row in [
        ("P1", "AZ-900 Fundamentals Reviewer.pdf", "Most fundamentals MC questions"),
        ("P2", "AZ900_Practice_Exam_With_Answers.pdf", "All True/False + several MC + scenarios"),
        ("P3", "RADOVAN_AZ-900_ Microsoft Azure Fundamentals Reviewer.pdf", "Domain & practice MC questions"),
    ]:
        ws.append(list(row))
    widths(ws, [10, 60, 40])
    for r in range(2, ws.max_row + 1):
        for c in range(1, 4):
            ws.cell(row=r, column=c).alignment = WRAP
            ws.cell(row=r, column=c).border = BORDER

    wb.save("AZ-900_Reference.xlsx")

# Original full use-case answers (verbatim from PDFs) - kept for study/reference
USECASES = [
 ("Networking", "Private dedicated link to Azure with no public internet", "ExpressRoute", "P2 Q7"),
 ("Networking", "Connect two Azure VNets", "Virtual network peering", "P2 Q7"),
 ("Networking", "Encrypted tunnel to Azure over the public internet", "VPN gateway", "P2 Q7"),
 ("Networking", "On-prem VPN device definition resource", "Local Network Gateway", "P2 Q9"),
 ("Security", "Password + one-time passcode = ?", "Multi-Factor Authentication (MFA)", "P2 Q11"),
 ("Shared Responsibility", "Two provider responsibilities in IaaS", "Physical hardware and physical datacenter security", "P2 Q12"),
 ("Compute", "Full Windows desktop from the cloud", "Azure Virtual Desktop", "P2 Q13"),
 ("Cloud Benefits", "Access to resources during a service failure", "High availability", "P2 Q14"),
 ("Shared Responsibility", "What the customer provides in SaaS", "Data and user access (accounts/identities)", "P2 Q15"),
 ("Shared Responsibility", "Shared responsibility in SaaS", "Identity and access management", "P2 Q16"),
 ("Service Models", "Lift-and-shift migration model", "IaaS (Infrastructure as a Service)", "P2 Q18"),
 ("Cloud Benefits", "Recover from catastrophic failure", "Disaster recovery", "P2 Q19"),
 ("Cloud Benefits", "Deploy resources close to users", "Geo-distribution", "P2 Q19"),
 ("Cloud Benefits", "Adjust resources to demand", "Scalability", "P2 Q19"),
 ("Cloud Benefits", "Deploy Azure resources close to users", "Geo-distribution (Azure Regions)", "P2 Q21"),
 ("Cost", "Plan when migrating a public website", "Pay monthly usage costs (OpEx)", "P2 Q22"),
 ("Governance", "Restrict resource creation to specific regions", "Azure Policy", "P2 Q25"),
 ("Management", "ARM template format", "JSON", "P2 Q28"),
 ("Governance", "Assign to every Azure resource", "Tags", "P2 Q31"),
 ("Management", "Identify unused VMs", "Azure Advisor", "P2 Q32"),
 ("Management", "Monitor health of Azure services", "Azure Service Health", "P2 Q33"),
 ("Security", "Show security recommendations", "Microsoft Defender for Cloud", "P2 Q33"),
 ("Governance", "Prevent accidental deletion", "Delete lock (CanNotDelete)", "P2 Q34"),
 ("Hybrid", "Extend compliance/monitoring to hybrid/multicloud", "Azure Arc", "P2 Q35"),
 ("Management", "Order to deploy 5 VMs from an existing VM", "Generalize VM, create image, create ARM template, deploy 5 VMs", "P2 Q36"),
 ("Cloud Models", "Most flexible cloud model", "Hybrid", "P3 D1 Q1"),
 ("Cost", "Pay only as resources are used", "OpEx / Consumption-based", "P3 D1 Q2"),
 ("Cloud Benefits", "Automatic addition of resources on demand", "Elasticity", "P3 D1 Q3"),
 ("Service Models", "Microsoft 365 Online category", "SaaS", "P3 D1 Q4"),
 ("Cloud Models", "Total control over security and resources", "Private", "P3 D1 Q5"),
 ("Architecture", "Minimum availability zones in a region", "3", "P3 D2 Q1"),
 ("Storage", "Tier for 180+ days, rarely accessed", "Archive", "P3 D2 Q2"),
 ("Networking", "Layer 7 load balancer for web apps", "Azure Application Gateway", "P3 D2 Q3"),
 ("Architecture", "Resource group deleted -> resources?", "They are deleted", "P3 D2 Q4"),
 ("Storage", "Physical tool to transfer ~40 TB to Azure", "Azure Data Box", "P3 D2 Q5"),
 ("Cost", "Largest VM discount, reclaimable anytime", "Spot Pricing / Spot Instances", "P3 D3 Q1"),
 ("Compliance", "Microsoft audit reports / compliance site", "Service Trust Portal", "P3 D3 Q2"),
 ("Governance", "Collection of Azure policies", "An Initiative", "P3 D3 Q3"),
 ("Shared Responsibility", "Provider always responsible for", "The physical network / physical datacenters", "P3 D3 Q4"),
 ("Management", "Recommends cost/security/reliability/perf", "Azure Advisor", "P3 D3 Q5"),
 ("Cloud Models", "AD on-prem, DB in cloud - which model", "Hybrid cloud", "P3 PT Q1"),
 ("Service Models", "Dev environment without managing OS", "PaaS (Platform as a Service)", "P3 PT Q2"),
 ("Networking", "Reach datacenters without public internet", "ExpressRoute", "P3 PT Q3"),
 ("Governance", "Restrict deployment regions", "Azure Policy", "P3 PT Q4"),
 ("Storage", "16 nines, copies to secondary region", "Geo-Redundant Storage (GRS)", "P3 PT Q5"),
 ("Identity", "Username + password sign-in process", "Authentication", "P3 PT Q6"),
 ("Architecture", "Permissions when moved to new resource group", "Inherits the new resource group's permissions", "P3 PT Q7"),
 ("Management", "ARM template file format", "JSON (Bicep also supported)", "P3 PT Q8"),
 ("Cost", "Cheapest VM, can be shut down anytime", "Spot pricing / Spot Instances", "P3 PT Q9"),
 ("Storage", "Two factors for disk pricing", "Quality (SSD vs HDD) and size", "P3 PT Q10"),
 ("Cloud Models", "Total control of hardware for single-org app", "Private cloud", "P3 PT Q11"),
 ("Management", "Check first for Azure-wide outage", "Azure Service Health", "P3 PT Q12"),
 ("Networking", "Isolate classroom traffic within one VNet", "Virtual subnets", "P3 PT Q13"),
 ("Security", "Minimum tier for Microsoft Defender for Cloud", "Free", "P3 PT Q15"),
 ("Architecture", "Basic building block of Azure", "A resource", "P1 p2"),
 ("Architecture", "Resource groups per resource", "Exactly one", "P1 p2"),
 ("Architecture", "Can resource groups be nested?", "No", "P1 p2"),
 ("Architecture", "Can a resource group be renamed?", "No", "P1 p2"),
 ("Architecture", "Management groups per directory / depth", "10,000 management groups, up to six levels", "P1 p2"),
 ("Architecture", "Two subscription boundary types", "Billing and Access control", "P1 p2"),
 ("Compute", "Container orchestration service", "Azure Kubernetes Service (AKS)", "P1 p3"),
 ("Compute", "Event-driven serverless code", "Azure Functions", "P1 p3"),
 ("Compute", "Azure Functions triggers", "HTTP requests, timers, and Azure Service Messages", "P1 p3"),
 ("Compute", "Four Azure App Service types", "Web Apps, API Apps, WebJobs, Mobile Apps", "P1 p3"),
 ("Compute", "VM Scale Set purpose", "Group of identical, load-balanced VMs that scale", "P1 p2"),
 ("Compute", "Availability sets group VMs by", "Update domains and fault domains", "P1 p2"),
 ("Networking", "NSG purpose", "Inbound/outbound rules to allow or block traffic", "P1 p3"),
 ("Networking", "ExpressRoute 100 Gbps to backbone", "ExpressRoute Direct", "P1 p4"),
 ("Networking", "VPN gateway connects via", "The public Internet (encrypted)", "P1 p4"),
 ("Networking", "Default VPN HA option", "Active/Standby", "P1 p4"),
 ("Storage", "Storage account naming rules", "3-24 lowercase letters/numbers, globally unique", "P1 p4"),
 ("Storage", "Default storage account type", "Standard general-purpose v2", "P1 p4"),
 ("Storage", "LRS copies", "3 copies in one datacenter (11 nines)", "P1 p4"),
 ("Storage", "ZRS copies", "3 copies across availability zones (12 nines)", "P1 p4"),
 ("Storage", "GRS copies", "3 local + 3 in a paired region (16 nines)", "P1 p4"),
 ("Storage", "RA-GRS adds", "Read access to the secondary region", "P1 p4"),
 ("Storage", "Unstructured data store", "Azure Blob", "P1 p4"),
 ("Storage", "NoSQL store", "Azure Tables", "P1 p4"),
 ("Storage", "Messaging queues", "Azure Queues", "P1 p4"),
 ("Storage", "Block-level VM disks", "Azure Disks", "P1 p4"),
 ("Storage", "Cloud file shares (SMB/NFS)", "Azure Files", "P1 p4"),
 ("Storage", "Data Box max capacity", "80 terabytes", "P1 p5"),
 ("Storage", "Data Box disk wipe standard", "NIST 800-88r1", "P1 p5"),
 ("Identity", "Microsoft Entra ID", "Cloud-based identity and access management", "P1 p5"),
 ("Identity", "Bridge on-prem AD to Entra ID", "Microsoft Entra Connect", "P1 p5"),
 ("Identity", "Four authentication methods", "Passwords, SSO, MFA, Passwordless", "P1 p6"),
 ("Identity", "Three passwordless options", "Windows Hello, Microsoft Authenticator, FIDO2 keys", "P1 p6"),
 ("Security", "Defender for Cloud three needs", "Continuously assess, secure, and defend", "P1 p6"),
 ("Security", "Seven Defense in Depth layers", "Physical; Identity/Access; Perimeter; Network; Compute; Application; Data", "P1 p6"),
 ("Governance", "Azure Policy effects", "Append, Audit, AuditIfNotExists, DeployIfNotExists, Deny, Disabled", "P1 p6"),
 ("Governance", "Azure Initiative", "A collection of Azure Policy conditions", "P1 p7"),
 ("Governance", "Resource tag parts", "A name and a value", "P1 p7"),
 ("Governance", "ReadOnly lock", "Read only - no modify or delete", "P1 p7"),
 ("Management", "ARM definition", "The service all Azure resource actions go through", "P1 p7"),
 ("Management", "ARM template languages", "JSON and Bicep", "P1 p7"),
 ("Security", "Key Vault purpose", "Store secrets, encryption keys, and certificates", "P1 p7"),
 ("Management", "Five Advisor categories", "Reliability, Security, Performance, Operational Excellence, Cost", "P1 p7"),
 ("Management", "Three Service Health components", "Azure Status, Service Health, Resource Health", "P1 p7"),
 ("Cost", "Pricing Calculator", "Estimates cost of future Azure deployments", "P1 p6"),
 ("Cost", "TCO Calculator", "Compares on-premises expenses to Azure", "P1 p6"),
 ("Cost", "Three cost alert types", "Budget, Credit, and Department spending quota alerts", "P1 p6"),
 ("Geography", "Region pair distance", "482 km / 300 miles", "P1 p2"),
 ("Geography", "Services not region-locked", "Microsoft Entra ID, Traffic Manager, Azure DNS", "P1 p2"),
 ("Cost", "99% uptime downtime per month", "7.2 hours per month", "P1 p1"),
 ("Security", "Zero Trust assumption", "A breach is happening; verify by authentication", "P3 3.1"),
 ("Security", "DDoS Protection tiers", "Basic (free) and Standard (paid)", "P3 3.5"),
 ("Support", "24x7 phone and email plan", "Standard", "P3 3.9"),
 ("SLA", "RPO definition", "Maximum acceptable data loss, measured in time", "P3 3.9"),
 ("SLA", "RTO definition", "How long it takes to restore service", "P3 3.9"),
 ("Lifecycle", "Feature lifecycle order", "Private Preview -> Public Preview -> General Availability", "P3 3.9"),
 ("Cost", "Azure Hybrid Benefit", "Reuse existing on-premises licenses in Azure", "P3 3.8"),
]

REDUNDANT = [
 ("ARM templates use JSON", "P2 Q28; P3 PT Q8; P1 p7", "Keep ONE (e.g., P2 Q28). Same fact 3x."),
 ("ExpressRoute = private, no public internet", "P2 Q7; P2 Q13; P3 PT Q3; P1 p3-4", "Keep ONE. Appears 4-5x."),
 ("VPN gateway = encrypted over public internet", "P2 Q7; P1 p4; P3 2.4", "Keep one."),
 ("Spot = cheapest, interruptible VMs", "P3 D3 Q1; P3 PT Q9; P1 p8", "Keep ONE; asked twice in P3."),
 ("Azure Advisor = recommendations / unused VMs", "P2 Q32; P3 D3 Q5; P1 p7", "Two angles; keep one."),
 ("Hybrid = most flexible / on-prem + cloud", "P3 D1 Q1; P3 PT Q1", "Keep one."),
 ("Private cloud = total control", "P3 D1 Q5; P3 PT Q11", "Keep one."),
 ("Azure Policy restricts allowed regions", "P2 Q25; P3 PT Q4; P3 3.6", "Keep one."),
 ("Deleting a resource group deletes its resources", "P3 D2 Q4; P1 p2", "Keep the question version."),
 ("Moved resource inherits new RG permissions", "P3 PT Q7; P1 p2", "Keep one."),
 ("Availability Zones - minimum of 3", "P3 D2 Q1; P1 p2", "Keep one."),
 ("Archive tier = 180+ days, rare access", "P3 D2 Q2; P1 p4", "Keep one."),
 ("GRS = 16 nines, secondary region", "P3 PT Q5; P1 p4", "Keep one."),
 ("Azure Data Box = physical bulk transfer", "P3 D2 Q5; P1 p5", "Keep one."),
 ("MFA = two DIFFERENT factor types", "P2 Q11; P3 PT Q14; P1 p6", "One definition + one MC is fine; both relate."),
 ("Authentication = proving identity", "P3 PT Q6; P2 Q5 stmt", "Mild overlap; T/F vs MC."),
 ("Service Health = check first for outage", "P2 Q33; P3 PT Q12; P1 p7", "Keep one."),
 ("Defender for Cloud = recommendations / Free tier", "P2 Q33; P3 PT Q15", "Different angles; related."),
 ("Tags = name/value metadata", "P2 Q30; P2 Q31; P1 p7", "Overlapping; trim."),
 ("Initiative = group of policies", "P3 D3 Q3; P1 p7", "Keep one."),
 ("Provider always owns physical layer", "P2 Q12; P3 D3 Q4; P1 p1", "Keep one."),
 ("Lift-and-shift => IaaS", "P2 Q18; P1 p1", "Keep one."),
 ("Disk pricing = quality + size", "P3 PT Q10; P3 2.5", "Keep one."),
 ("Elasticity = automatic scaling", "P3 D1 Q3; P1 p1", "Keep one."),
 ("SaaS example = Microsoft 365 Online", "P3 D1 Q4; P1 p1", "Keep one."),
]

mc_err = build_mc_file()
tf_err = build_tf_file()
build_reference_file()

print("Multiple-choice questions :", len(MC))
print("True/False questions      :", len(TF))
print("TOTAL importable items    :", len(MC) + len(TF))
print("Reference use-case rows   :", len(USECASES))
print()
errs = mc_err + tf_err
if errs:
    print("!!! CHARACTER-LIMIT VIOLATIONS (must fix):")
    for e in errs:
        print("  ", e)
else:
    print("OK - all questions <=95 chars and all answers <=60 chars.")
